"""
Interface de Configuração - Sistema Genérico para VTEX
"""
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QGroupBox, QComboBox, QCheckBox, QTabWidget, QWidget, QMessageBox,
    QFormLayout, QSpinBox, QColorDialog, QFileDialog
)
from PySide6.QtGui import QFont, QColor
from PySide6.QtCore import Qt, Signal
from config_manager import ConfigManager
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ConfigDialog(QDialog):
    """Dialog de configuração da empresa"""
    
    config_updated = Signal()  # Sinal emitido quando configuração é atualizada
    empresa_updated = Signal(str)  # Sinal emitido quando nome da empresa é atualizado
    lojas_updated = Signal()  # Sinal emitido quando lojas são alteradas
    tokens_updated = Signal()  # Sinal emitido quando tokens são alterados
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.config_manager = ConfigManager()
        self.parent = parent
        self.setWindowTitle("Configuração da Empresa")
        self.setModal(True)
        self.resize(800, 600)
        self.init_ui()
        self.load_config()
        
        # Sinais de atualização ao vivo removidos
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Criar abas
        self.tab_widget = QTabWidget()
        layout.addWidget(self.tab_widget)
        
        # Aba de Informações da Empresa
        self.empresa_tab = self.create_empresa_tab()
        self.tab_widget.addTab(self.empresa_tab, "Empresa")
        
        # Aba de Lojas
        self.lojas_tab = self.create_lojas_tab()
        self.tab_widget.addTab(self.lojas_tab, "Lojas")
        
        # Sem abas de cores e splash - apenas empresa e lojas
        
        # Aba de Configurações
        self.config_tab = self.create_config_tab()
        self.tab_widget.addTab(self.config_tab, "Configurações")
        
        # Botões
        buttons_layout = QHBoxLayout()
        
        self.salvar_btn = QPushButton("Salvar")
        self.salvar_btn.clicked.connect(self.salvar_config)
        
        self.cancelar_btn = QPushButton("Cancelar")
        self.cancelar_btn.clicked.connect(self.reject)
        
        self.importar_btn = QPushButton("Importar Config")
        self.importar_btn.clicked.connect(self.importar_config)
        
        self.exportar_btn = QPushButton("Exportar Config")
        self.exportar_btn.clicked.connect(self.exportar_config)
        
        buttons_layout.addWidget(self.importar_btn)
        buttons_layout.addWidget(self.exportar_btn)
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.salvar_btn)
        buttons_layout.addWidget(self.cancelar_btn)
        
        layout.addLayout(buttons_layout)
    
    def create_empresa_tab(self):
        """Cria aba de informações da empresa"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Grupo de informações da empresa
        empresa_group = QGroupBox("Informações da Empresa")
        empresa_layout = QFormLayout(empresa_group)
        
        self.empresa_nome = QLineEdit()
        self.empresa_nome.setPlaceholderText("Nome da empresa")
        empresa_layout.addRow("Nome:", self.empresa_nome)
        
        self.conta_principal = QLineEdit()
        self.conta_principal.setPlaceholderText("trackfield")
        self.conta_principal.setReadOnly(True)  # Não editável - definida pela lista de lojas
        self.conta_principal.setStyleSheet("background-color: #f0f0f0; color: #666;")
        empresa_layout.addRow("Conta Principal:", self.conta_principal)
        
        self.app_key = QLineEdit()
        self.app_key.setPlaceholderText("App Key da VTEX")
        self.app_key.setEchoMode(QLineEdit.Password)
        empresa_layout.addRow("App Key:", self.app_key)
        
        self.app_token = QLineEdit()
        self.app_token.setPlaceholderText("App Token da VTEX")
        self.app_token.setEchoMode(QLineEdit.Password)
        empresa_layout.addRow("App Token:", self.app_token)
        
        # Removido user_token - não é mais necessário
        
        layout.addWidget(empresa_group)
        layout.addStretch()
        
        return tab
    
    def create_lojas_tab(self):
        """Cria aba de gerenciamento de lojas"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Botões de ação
        buttons_layout = QHBoxLayout()
        
        self.adicionar_loja_btn = QPushButton("Adicionar Loja")
        self.adicionar_loja_btn.clicked.connect(self.adicionar_loja)
        
        self.editar_loja_btn = QPushButton("Editar Loja")
        self.editar_loja_btn.clicked.connect(self.editar_loja)
        
        self.remover_loja_btn = QPushButton("Remover Loja")
        self.remover_loja_btn.clicked.connect(self.remover_loja)
        
        # Botões de Excel
        self.exportar_excel_btn = QPushButton("📊 Exportar Excel")
        self.exportar_excel_btn.clicked.connect(self.exportar_lojas_excel)
        
        self.importar_excel_btn = QPushButton("📥 Importar Excel")
        self.importar_excel_btn.clicked.connect(self.importar_lojas_excel)
        
        buttons_layout.addWidget(self.adicionar_loja_btn)
        buttons_layout.addWidget(self.editar_loja_btn)
        buttons_layout.addWidget(self.remover_loja_btn)
        buttons_layout.addWidget(self.exportar_excel_btn)
        buttons_layout.addWidget(self.importar_excel_btn)
        buttons_layout.addStretch()
        
        layout.addLayout(buttons_layout)
        
        # Tabela de lojas simplificada
        self.lojas_table = QTableWidget()
        self.lojas_table.setColumnCount(6)
        self.lojas_table.setHorizontalHeaderLabels([
            "ID", "Nome", "Tipo", "Nacional", "Propriedade", "Conta Principal"
        ])
        
        # Configurar tabela
        header = self.lojas_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID
        header.setSectionResizeMode(1, QHeaderView.Stretch)           # Nome
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # Tipo
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Nacional
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Propriedade
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # Conta Principal
        
        self.lojas_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.lojas_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        
        layout.addWidget(self.lojas_table)
        
        return tab
    
    # Método de cores removido - cores fixas
    
    # Método de splash removido - splash fixa
    
    def create_config_tab(self):
        """Cria aba de configurações gerais"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        config_group = QGroupBox("Configurações Gerais")
        config_layout = QFormLayout(config_group)
        
        self.sku_padrao = QLineEdit()
        self.sku_padrao.setPlaceholderText("149718")
        config_layout.addRow("SKU Padrão:", self.sku_padrao)
        
        self.max_skus = QSpinBox()
        self.max_skus.setRange(1, 20)
        self.max_skus.setValue(5)
        config_layout.addRow("Máx SKUs Recentes:", self.max_skus)
        
        self.max_workers = QSpinBox()
        self.max_workers.setRange(1, 50)
        self.max_workers.setValue(20)
        config_layout.addRow("Máx Workers:", self.max_workers)
        
        self.timeout = QSpinBox()
        self.timeout.setRange(5, 60)
        self.timeout.setValue(10)
        config_layout.addRow("Timeout (segundos):", self.timeout)
        
        layout.addWidget(config_group)
        layout.addStretch()
        
        return tab
    
    def load_config(self):
        """Carrega configurações atuais"""
        # Carregar informações da empresa
        empresa = self.config_manager.get_empresa_info()
        self.empresa_nome.setText(empresa.get("nome", ""))
        self.conta_principal.setText(empresa.get("conta_principal", ""))
        self.app_key.setText(empresa.get("app_key", ""))
        self.app_token.setText(empresa.get("app_token", ""))
        # Removido user_token - não é mais necessário
        
        # Carregar lojas
        self.load_lojas_table()
        
        # Atualizar conta principal baseada na lista de lojas
        self._atualizar_conta_principal_display()
        
        # Cores e splash removidas - mantidas fixas
        
        # Carregar configurações
        config = self.config_manager.get_configuracoes()
        self.sku_padrao.setText(config.get("sku_padrao", ""))
        self.max_skus.setValue(config.get("max_skus_recentes", 5))
        self.max_workers.setValue(config.get("max_workers", 20))
        self.timeout.setValue(config.get("timeout_requests", 10))
    
    def load_lojas_table(self):
        """Carrega tabela de lojas simplificada"""
        lojas = self.config_manager.get_lojas()
        self.lojas_table.setRowCount(len(lojas))
        
        for i, loja in enumerate(lojas):
            self.lojas_table.setItem(i, 0, QTableWidgetItem(loja.get("id", "")))
            self.lojas_table.setItem(i, 1, QTableWidgetItem(loja.get("nome", "")))
            self.lojas_table.setItem(i, 2, QTableWidgetItem(loja.get("tipo", "")))
            
            # Checkbox para nacional
            nacional_item = QTableWidgetItem()
            nacional_item.setFlags(nacional_item.flags() | Qt.ItemIsUserCheckable)
            nacional_item.setCheckState(Qt.Checked if loja.get("nacional", False) else Qt.Unchecked)
            self.lojas_table.setItem(i, 3, nacional_item)
            
            # Propriedade (Franquia/Própria)
            self.lojas_table.setItem(i, 4, QTableWidgetItem(loja.get("propriedade", "")))
            
            # Checkbox para conta principal
            principal_item = QTableWidgetItem()
            principal_item.setFlags(principal_item.flags() | Qt.ItemIsUserCheckable)
            principal_item.setCheckState(Qt.Checked if loja.get("conta_principal", False) else Qt.Unchecked)
            self.lojas_table.setItem(i, 5, principal_item)
    
    # Método de escolher cor removido - cores fixas
    
    # Método de escolher logo removido - splash fixa
    
    def _desmarcar_outras_contas_principais(self):
        """Desmarca outras lojas como conta principal"""
        lojas = self.config_manager.get_lojas()
        for loja in lojas:
            if loja.get("conta_principal", False):
                loja["conta_principal"] = False
        self.config_manager.save_config()
    
    def _atualizar_conta_principal(self):
        """Atualiza a conta principal baseada na loja marcada"""
        lojas = self.config_manager.get_lojas()
        conta_principal_id = None
        
        # Encontrar qual loja está marcada como conta principal
        for loja in lojas:
            if loja.get("conta_principal", False):
                conta_principal_id = loja.get("id")
                break
        
        # Atualizar o campo conta_principal na seção empresa
        if conta_principal_id:
            self.config_manager.config["empresa"]["conta_principal"] = conta_principal_id
        else:
            # Se nenhuma loja estiver marcada, usar a primeira loja
            if lojas:
                self.config_manager.config["empresa"]["conta_principal"] = lojas[0].get("id", "")
    
    def _atualizar_conta_principal_display(self):
        """Atualiza o display da conta principal baseado na lista de lojas"""
        lojas = self.config_manager.get_lojas()
        conta_principal_id = None
        
        # Encontrar qual loja está marcada como conta principal
        for loja in lojas:
            if loja.get("conta_principal", False):
                conta_principal_id = loja.get("id")
                break
        
        # Atualizar o campo na interface
        if conta_principal_id:
            self.conta_principal.setText(conta_principal_id)
        else:
            # Se nenhuma loja estiver marcada, usar a primeira loja
            if lojas:
                self.conta_principal.setText(lojas[0].get("id", ""))
            else:
                self.conta_principal.setText("Nenhuma loja definida")
    
    def adicionar_loja(self):
        """Adiciona nova loja"""
        dialog = LojaDialog(self)
        if dialog.exec() == QDialog.Accepted:
            loja_data = dialog.get_loja_data()
            
            # Verificar se está marcando como conta principal
            if loja_data.get("conta_principal", False):
                # Desmarcar outras lojas como conta principal
                self._desmarcar_outras_contas_principais()
            
            if self.config_manager.add_loja(loja_data):
                self.load_lojas_table()
                self._atualizar_conta_principal_display()
                QMessageBox.information(self, "Sucesso", "Loja adicionada com sucesso!")
            else:
                QMessageBox.warning(self, "Erro", "Erro ao adicionar loja ou loja já existe!")
    
    def editar_loja(self):
        """Edita loja selecionada"""
        current_row = self.lojas_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Aviso", "Selecione uma loja para editar!")
            return
        
        loja_id = self.lojas_table.item(current_row, 0).text()
        lojas = self.config_manager.get_lojas()
        loja_data = None
        
        for loja in lojas:
            if loja.get("id") == loja_id:
                loja_data = loja
                break
        
        if loja_data:
            dialog = LojaDialog(self, loja_data)
            if dialog.exec() == QDialog.Accepted:
                nova_loja_data = dialog.get_loja_data()
                
                # Verificar se está marcando como conta principal
                if nova_loja_data.get("conta_principal", False):
                    # Desmarcar outras lojas como conta principal
                    self._desmarcar_outras_contas_principais()
                
                if self.config_manager.update_loja(loja_id, nova_loja_data):
                    self.load_lojas_table()
                    self._atualizar_conta_principal_display()
                    QMessageBox.information(self, "Sucesso", "Loja atualizada com sucesso!")
                else:
                    QMessageBox.warning(self, "Erro", "Erro ao atualizar loja!")
    
    def remover_loja(self):
        """Remove loja selecionada"""
        current_row = self.lojas_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Aviso", "Selecione uma loja para remover!")
            return
        
        loja_id = self.lojas_table.item(current_row, 0).text()
        loja_nome = self.lojas_table.item(current_row, 1).text()
        
        reply = QMessageBox.question(
            self, "Confirmar Remoção",
            f"Tem certeza que deseja remover a loja '{loja_nome}'?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            if self.config_manager.remove_loja(loja_id):
                self.load_lojas_table()
                QMessageBox.information(self, "Sucesso", "Loja removida com sucesso!")
            else:
                QMessageBox.warning(self, "Erro", "Erro ao remover loja!")
    
    def salvar_config(self):
        """Salva todas as configurações"""
        try:
            # Salvar informações da empresa
            empresa_data = {
                "nome": self.empresa_nome.text(),
                "conta_principal": self.conta_principal.text(),
                "app_key": self.app_key.text(),
                "app_token": self.app_token.text(),
                # Removido user_token - não é mais necessário
            }
            self.config_manager.update_empresa_info(empresa_data)
            
            # Cores e splash removidas - mantidas fixas
            
            # Atualizar conta principal baseada na loja marcada
            self._atualizar_conta_principal()
            
            # Salvar configurações gerais
            config_data = {
                "sku_padrao": self.sku_padrao.text(),
                "max_skus_recentes": self.max_skus.value(),
                "max_workers": self.max_workers.value(),
                "timeout_requests": self.timeout.value()
            }
            self.config_manager.config["configuracoes"] = config_data
            
            # Salvar arquivo
            if self.config_manager.save_config():
                QMessageBox.information(self, "Sucesso", "Configurações salvas com sucesso!")
                
                
                self.accept()
            else:
                QMessageBox.warning(self, "Erro", "Erro ao salvar configurações!")
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar configurações: {str(e)}")
    
    def exportar_lojas_excel(self):
        """Exporta lista de lojas para Excel"""
        try:
            # Solicitar local para salvar
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Exportar Lojas para Excel", 
                "lojas.xlsx", 
                "Arquivos Excel (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Lojas"
            
            # Cabeçalhos
            headers = ["ID", "Nome", "Tipo", "Nacional", "Propriedade", "Conta Principal"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Dados das lojas
            lojas = self.config_manager.config.get("lojas", [])
            for row, loja in enumerate(lojas, 2):
                ws.cell(row=row, column=1, value=loja.get("id", ""))
                ws.cell(row=row, column=2, value=loja.get("nome", ""))
                ws.cell(row=row, column=3, value=loja.get("tipo", ""))
                ws.cell(row=row, column=4, value="Sim" if loja.get("nacional", False) else "Não")
                ws.cell(row=row, column=5, value=loja.get("propriedade", ""))
                ws.cell(row=row, column=6, value="Sim" if loja.get("conta_principal", False) else "Não")
            
            # Ajustar largura das colunas
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 15
            
            # Salvar arquivo
            wb.save(file_path)
            QMessageBox.information(self, "Sucesso", f"Lojas exportadas para:\n{file_path}")
            
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao exportar Excel: {str(e)}")
    
    def importar_lojas_excel(self):
        """Importa lista de lojas do Excel"""
        try:
            # Solicitar arquivo
            file_path, _ = QFileDialog.getOpenFileName(
                self, 
                "Importar Lojas do Excel", 
                "", 
                "Arquivos Excel (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Carregar workbook
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Verificar cabeçalhos
            headers = [ws.cell(row=1, column=col).value for col in range(1, 7)]
            expected_headers = ["ID", "Nome", "Tipo", "Nacional", "Propriedade", "Conta Principal"]
            
            if headers != expected_headers:
                QMessageBox.warning(self, "Erro", 
                    "Formato de arquivo inválido!\n"
                    "Cabeçalhos esperados: ID, Nome, Tipo, Nacional, Propriedade, Conta Principal")
                return
            
            # Ler dados
            lojas_importadas = []
            for row in range(2, ws.max_row + 1):
                if not ws.cell(row=row, column=1).value:  # Pular linhas vazias
                    continue
                    
                loja = {
                    "id": str(ws.cell(row=row, column=1).value or ""),
                    "nome": str(ws.cell(row=row, column=2).value or ""),
                    "tipo": str(ws.cell(row=row, column=3).value or "Local"),
                    "nacional": str(ws.cell(row=row, column=4).value or "").lower() in ["sim", "s", "yes", "y", "1", "true"],
                    "propriedade": str(ws.cell(row=row, column=5).value or "Própria"),
                    "conta_principal": str(ws.cell(row=row, column=6).value or "").lower() in ["sim", "s", "yes", "y", "1", "true"]
                }
                
                if loja["id"] and loja["nome"]:  # Validar dados obrigatórios
                    lojas_importadas.append(loja)
            
            if not lojas_importadas:
                QMessageBox.warning(self, "Aviso", "Nenhuma loja válida encontrada no arquivo!")
                return
            
            # Confirmar importação
            reply = QMessageBox.question(
                self, "Confirmar Importação",
                f"Importar {len(lojas_importadas)} lojas?\n"
                "Isso substituirá a lista atual de lojas.",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                # Substituir lojas
                self.config_manager.config["lojas"] = lojas_importadas
                
                # Atualizar conta principal
                self._atualizar_conta_principal()
                
                # Recarregar tabela
                self.load_lojas_table()
                self._atualizar_conta_principal_display()
                
                QMessageBox.information(self, "Sucesso", 
                    f"{len(lojas_importadas)} lojas importadas com sucesso!")
            
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao importar Excel: {str(e)}")
    
    def importar_config(self):
        """Importa configurações de arquivo"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Importar Configurações", "", "JSON Files (*.json)"
        )
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                self.config_manager.config = config
                self.load_config()
                QMessageBox.information(self, "Sucesso", "Configurações importadas com sucesso!")
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao importar configurações: {str(e)}")
    
    def exportar_config(self):
        """Exporta configurações para arquivo"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Exportar Configurações", "empresa_config.json", "JSON Files (*.json)"
        )
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(self.config_manager.config, f, indent=2, ensure_ascii=False)
                QMessageBox.information(self, "Sucesso", "Configurações exportadas com sucesso!")
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao exportar configurações: {str(e)}")


class LojaDialog(QDialog):
    """Dialog para adicionar/editar loja"""
    
    def __init__(self, parent=None, loja_data=None):
        super().__init__(parent)
        self.loja_data = loja_data
        self.setWindowTitle("Adicionar/Editar Loja" if loja_data else "Adicionar Loja")
        self.setModal(True)
        self.resize(400, 300)
        self.init_ui()
        
        if loja_data:
            self.load_loja_data()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        form_layout = QFormLayout()
        
        # ID da Loja
        self.loja_id = QLineEdit()
        self.loja_id.setPlaceholderText("trackfield")
        form_layout.addRow("ID da Loja:", self.loja_id)
        
        # Nome da Loja
        self.loja_nome = QLineEdit()
        self.loja_nome.setPlaceholderText("Nome da Loja")
        form_layout.addRow("Nome:", self.loja_nome)
        
        # Tipo (Nacional/Local)
        self.loja_tipo = QComboBox()
        self.loja_tipo.addItems(["Nacional", "Local"])
        form_layout.addRow("Tipo:", self.loja_tipo)
        
        # Propriedade (Franquia/Própria)
        self.propriedade = QComboBox()
        self.propriedade.addItems(["Franquia", "Própria"])
        form_layout.addRow("Propriedade:", self.propriedade)
        
        # Conta Principal (checkbox)
        self.conta_principal = QCheckBox("Esta é a conta principal")
        form_layout.addRow("", self.conta_principal)
        
        layout.addLayout(form_layout)
        
        # Botões
        buttons_layout = QHBoxLayout()
        
        self.salvar_btn = QPushButton("Salvar")
        self.salvar_btn.clicked.connect(self.accept)
        
        self.cancelar_btn = QPushButton("Cancelar")
        self.cancelar_btn.clicked.connect(self.reject)
        
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.salvar_btn)
        buttons_layout.addWidget(self.cancelar_btn)
        
        layout.addLayout(buttons_layout)
    
    def load_loja_data(self):
        """Carrega dados da loja para edição"""
        if self.loja_data:
            self.loja_id.setText(self.loja_data.get("id", ""))
            self.loja_nome.setText(self.loja_data.get("nome", ""))
            self.loja_tipo.setCurrentText(self.loja_data.get("tipo", "Local"))
            self.propriedade.setCurrentText(self.loja_data.get("propriedade", "Própria"))
            self.conta_principal.setChecked(self.loja_data.get("conta_principal", False))
    
    def get_loja_data(self):
        """Retorna dados da loja simplificados"""
        return {
            "id": self.loja_id.text(),
            "nome": self.loja_nome.text(),
            "tipo": self.loja_tipo.currentText(),
            "nacional": self.loja_tipo.currentText() == "Nacional",
            "propriedade": self.propriedade.currentText(),
            "conta_principal": self.conta_principal.isChecked()
        }
